from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import ApplicationBuilder, CommandHandler, CallbackQueryHandler, ContextTypes
import openpyxl
from openpyxl import Workbook
import os

TOKEN = "7519116906:AAFafU4Z00Cyww1RXVvvluLGn11HGjrc7GA"

STATS_FILE = 'uno_game_stats.xlsx'

# Якщо файл не існує, створюємо новий
if not os.path.exists(STATS_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(['Game ID', 'Player', 'Games Played', 'Wins', 'Cards Played'])  # заголовки
    wb.save(STATS_FILE)

# In-memory data (in production, replace with a database)
games = {}  # chat_id: game_data
user_settings = {}  # user_id: settings
user_stats = {}  # user_id: stats

# Helper functions (simplified)
def is_admin(user_id, chat_id):
    # In real use, check admin status properly
    return True

def get_game(chat_id):
    return games.get(chat_id)

def save_game_stats(player_id, game_id, stats):
    wb = openpyxl.load_workbook(STATS_FILE)
    ws = wb.active
    ws.append([game_id, player_id, stats['games'], stats['wins'], stats['cards']])
    wb.save(STATS_FILE)

async def send_private_message(user_id, text):
    try:
        await application.bot.send_message(chat_id=user_id, text=text)
    except:
        pass

# Command Handlers
async def new(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    user_id = update.effective_user.id
    if chat_id in games:
        await update.message.reply_text("Game already in progress.")
        return
    games[chat_id] = {
        'creator': user_id,
        'players': [user_id],
        'open': True,
        'turn_order': [],
        'direction': 'clockwise',
        'game_id': chat_id
    }
    await update.message.reply_text("New UNO game created! Use /join to participate.")

async def join(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    user_id = update.effective_user.id
    game = get_game(chat_id)
    if not game:
        await update.message.reply_text("No game to join. Use /new to create one.")
        return
    if not game['open']:
        await update.message.reply_text("Game lobby is closed.")
        return
    if user_id not in game['players']:
        game['players'].append(user_id)
        await update.message.reply_text(f"{update.effective_user.first_name} joined the game!")
    else:
        await update.message.reply_text("You already joined.")

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.message.chat_id
    game = get_game(chat_id)
    if chat_id in games and len(games[chat_id]["players"]) < 4:
        await update.message.reply_text("Not enough players to start the game, need at least 4 players to start.")
        return
        game['turn_order'] = list(game['players'])
    await update.message.reply_text(START_TEXT)
    if chat_id in games and len(games[chat_id]["players"]) > 4:
    await update.message.reply_text("Game started!")

    # Send instructions to each player privately
    for player_id in game['players']:
        await send_private_message(player_id,
    START_TEXT = "Follow these steps:"
            "1️⃣ Add this bot to a group"
            "2️⃣ In the group, start a new game with /new or join an already running game with /join"
            "3️⃣ After at least two players have joined, start the game with /start"
            "🃏 Type @Merciless_UnoBot into your chat box and hit space, or click the 'via @Merciless_UnoBot' text next to messages. "
            "You will see your cards (some greyed out), any extra options like drawing, to see the current game state write /turn_order. "
            "The greyed out cards are those you cannot play at the moment. Tap an option to execute the selected action."
            "Players can join the game at any time. To leave a game, use /leave. If a player takes more than 90 seconds to play, "
            "you can use /skip to skip that player. Use /notify_me to receive a private message when a new game is started."
            "⚙️ Other settings: /settings"
            "👑 Other commands (only game creator or admin):"
            "/close - Close lobby"
            "/open - Open lobby"
            "/kill - Terminate the game"
            "/kick - Select a player to kick by replying to him or her")

async def leave_game(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    user_id = update.effective_user.id
    game = get_game(chat_id)
    if game and user_id in game['players']:
        game['players'].remove(user_id)
        await update.message.reply_text("You left the game.")
    else:
        await update.message.reply_text("You're not in a game.")

async def turn_order(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    game = get_game(chat_id)
    if not game:
        await update.message.reply_text("No game in progress.")
        return

    # Формуємо повідомлення про поточного гравця та кількість карт
    player_status = []
    for player_id in game['players']:
        player_name = f"Player {game['players'].index(player_id) + 1}"  # Placeholder for player name
        player_status.append(f"{player_name} ({len(game['players'])} cards)")  # Replace with actual card count

    turn_order_text = " -> ".join(player_status)
    current_player = f"Current player: Player {game['turn_order'][0]} (@{game['turn_order'][0]})"  # Replace with actual player tag
    last_card = "💛4"  # Placeholder for last card played

    await update.message.reply_text(f"{current_player}\nLast card: {last_card}\nPlayers: {turn_order_text}")

async def save_game_stats(player_id, game_id, stats):
    wb = openpyxl.load_workbook(STATS_FILE)
    ws = wb.active
    ws.append([game_id, player_id, stats['games'], stats['wins'], stats['cards']])
    wb.save(STATS_FILE)

async def settings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await send_private_message(update.effective_user.id, "Settings (placeholder). Enable stats here.")

async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    stats = user_stats.get(user_id, {'games': 0, 'wins': 0, 'cards': 0})
    percent = (stats['wins'] / stats['games'] * 100) if stats['games'] > 0 else 0
    await update.message.reply_text(
        f"Games Played: {stats['games']}\nWins: {stats['wins']} ({percent:.2f}%)\nCards Played: {stats['cards']}"
    )

# Bot Setup
app = Application.builder().token(TOKEN).build()
app.job_queue.scheduler.configure(timezone=pytz.UTC)
application.add_handler(CommandHandler("new", new_game))
application.add_handler(CommandHandler("join", join_game))
application.add_handler(CommandHandler("start", start_game))
application.add_handler(CommandHandler("leave", leave_game))
application.add_handler(CommandHandler("turn_order", turn_order))
application.add_handler(CommandHandler("settings", settings))
application.add_handler(CommandHandler("stats", stats))

if __name__ == '__main__':
    application.run_polling()
